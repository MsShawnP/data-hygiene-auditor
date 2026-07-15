"""Generate a messy sample Excel file for testing the Data Hygiene Auditor."""
from pathlib import Path

from openpyxl import Workbook


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    headers = [
        "CustomerID", "FirstName", "LastName", "Email", "Phone",
        "JoinDate", "AccountBalance", "Status", "ZipCode", "Notes"
    ]
    ws.append(headers)

    # --- Row data with intentional problems ---
    rows = [
        # Clean-ish rows
        ["CUST-001", "Alice", "Johnson", "alice@example.com", "(555) 123-4567", "2023-01-15", "$1,250.00", "Active", "30301", "Preferred customer"],
        ["CUST-002", "Bob", "Smith", "bob.smith@example.com", "555-234-5678", "01/15/2023", "1250.00", "Active", "30302", ""],
        ["CUST-003", "Charlie", "Williams", "charlie.w@example.com", "5553456789", "Jan 15, 2023", "$2,500", "active", "30303", "VIP"],
        # Mixed date formats
        ["CUST-004", "Diana", "Brown", "diana.b@example.com", "(555) 456-7890", "2023-02-20", "$3,100.50", "Active", "30304", "Referred by Alice"],
        ["CUST-005", "Edward", "Davis", "edward@example.com", "555.567.8901", "02/20/2023", "3100.50", "Inactive", "30305", ""],
        ["CUST-006", "Fiona", "Garcia", "fiona.g@example.com", "+1-555-678-9012", "Feb 20, 2023", "$4,200.00", "ACTIVE", "30306", "Corporate account"],
        # Phantom duplicates (whitespace/case variations)
        ["CUST-007", "alice", "johnson", "alice@example.com", "(555) 123-4567", "2023-01-15", "$1,250.00", "Active", "30301", "Preferred customer"],
        ["CUST-008", " Alice ", " Johnson", "ALICE@EXAMPLE.COM", "(555) 123-4567", "2023-01-15", "$1,250.00", "Active", "30301", "Preferred customer"],
        ["CUST-009", "Bob", "Smith ", "bob.smith@example.com ", "555-234-5678", "01/15/2023", "1250.00", "Active", "30302", ""],
        # Suspiciously uniform / placeholder data
        ["CUST-010", "Test", "User", "test@test.com", "000-000-0000", "2023-01-01", "$0.00", "Active", "00000", "TEST"],
        ["CUST-011", "Test", "User", "test@test.com", "000-000-0000", "2023-01-01", "$0.00", "Active", "00000", "TEST"],
        ["CUST-012", "Test", "User", "test@test.com", "000-000-0000", "2023-01-01", "$0.00", "Active", "00000", "TEST"],
        ["CUST-013", "N/A", "N/A", "n/a", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A"],
        ["CUST-014", "TBD", "TBD", "tbd@tbd.com", "TBD", "TBD", "TBD", "TBD", "TBD", "TBD"],
        # Numbers stored as text, codes in name fields
        ["CUST-015", "REF-4421", "Martinez", "martinez@example.com", "(555) 789-0123", "2023-03-10", "five thousand", "1", "303-07", ""],
        ["CUST-016", "Grace", "ABC-CORP-2023", "grace@example.com", "not available", "March 10 2023", "$5,000", "Active", "30308", "See ticket #4421"],
        ["CUST-017", "Henry", "Lee", "henrylee", "(555) 890-1234", "2023/03/15", "$6,100.00", "Active", "30309", ""],
        # Missing data flood
        ["CUST-018", "", "", "", "", "", "", "", "", ""],
        ["CUST-019", None, None, None, None, None, None, None, None, None],
        ["CUST-020", "  ", "  ", "  ", "  ", "  ", "  ", "  ", "  ", "  "],
        # More mixed formats
        ["CUST-021", "Irene", "Wilson", "irene.w@example.com", "555 012 3456", "3/15/2023", "$7200", "Suspended", "30310", "Payment issue"],
        ["CUST-022", "Jack", "Taylor", "jack.t@example.com", "(555)0123456", "15-Mar-2023", "7,200.00", "Active", "30311", ""],
        ["CUST-023", "Karen", "Anderson", "karen.a@example.com", "1-555-123-4567", "2023.03.15", "$8,500.00", "active", "30312-1234", "Extended zip"],
        # More suspicious uniformity
        ["CUST-024", "John", "Doe", "john@doe.com", "555-555-5555", "2023-01-01", "$0.00", "Active", "12345", ""],
        ["CUST-025", "Jane", "Doe", "jane@doe.com", "555-555-5555", "2023-01-01", "$0.00", "Active", "12345", ""],
        ["CUST-026", "John", "Doe", "john@doe.com", "555-555-5555", "2023-01-01", "$0.00", "Active", "12345", ""],
        # Wrong-purpose fields
        ["1027", "Lisa", "Thomas", "lisa.t@example.com", "(555) 234-5678", "2023-04-01", "$9,100.00", "Active", "30313", ""],
        ["CUST-028", "Mike", "Jackson", "mike.j@example.com", "(555) 345-6789", "2023-04-05", "$10,250.00 USD", "Y", "30314", "Balance includes pending"],
        ["CUST-029", "Nancy", "White", "nancy w@example.com", "(555) 456 7890", "04-05-2023", "$-500.00", "Active", "3031", "Credit balance"],
        ["CUST-030", "Oscar", "Harris", "oscar@example.com", "(555) 567-8901", "2023-04-10", "$11,000.00", "Active", "30316", ""],
    ]

    for row in rows:
        ws.append(row)

    # Sheet 2: Orders (smaller, to show multi-sheet support)
    ws2 = wb.create_sheet("Orders")
    order_headers = ["OrderID", "CustomerID", "OrderDate", "Amount", "ShipDate", "Status"]
    ws2.append(order_headers)
    order_rows = [
        ["ORD-001", "CUST-001", "2023-06-01", "$150.00", "2023-06-03", "Shipped"],
        ["ORD-002", "CUST-001", "06/15/2023", "200", "06/17/2023", "Shipped"],
        ["ORD-003", "CUST-002", "Jun 20, 2023", "$175.50", "Jun 22 2023", "shipped"],
        ["ORD-004", "CUST-003", "2023-07-01", "$300.00", "2023-07-03", "Delivered"],
        ["ORD-005", "CUST-003", "2023-07-01", "$300.00", "2023-07-03", "Delivered"],
        ["ORD-006", "CUST-010", "2023-01-01", "$0.00", "2023-01-01", "Test"],
        ["ORD-007", "CUST-010", "2023-01-01", "$0.00", "2023-01-01", "Test"],
        ["ORD-008", "cust-003", "7/1/2023", "300", "7/3/2023", "delivered"],
        ["ORD-009", "CUST-004", "2023-08-10", "$450.00", "", "Pending"],
        ["ORD-010", "CUST-004", "2023/08/10", "$450", None, "PENDING"],
    ]
    for row in order_rows:
        ws2.append(row)

    return wb


def build_realistic_workbook():
    """A moderately messy, realistic membership export.

    Unlike ``build_workbook`` (a deliberate torture test), this file looks
    like a real small-business roster that has drifted: a handful of mixed
    date/phone/currency formats, two placeholder cells, one lowercase
    status, a little missing data, one exact double-entry, and one
    case-variant duplicate. Most rows are clean. It exists so the health
    score has something to discriminate against — a file that scores in the
    "needs attention" band rather than the "critical" band.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Members"

    headers = [
        "MemberID", "FirstName", "LastName", "Email", "Phone",
        "JoinDate", "MonthlyFee", "Status", "ZipCode", "Notes"
    ]
    ws.append(headers)

    rows = [
        # Clean core: ISO dates, (555) xxx-xxxx phones, $xx.00 fees
        ["M-1001", "Sarah", "Chen", "sarah.chen@example.com", "(555) 201-3345", "2024-01-08", "$49.00", "Active", "97201", "Founding member"],
        ["M-1002", "David", "Okafor", "d.okafor@example.com", "(555) 202-1190", "2024-01-12", "$49.00", "Active", "97202", ""],
        ["M-1003", "Maria", "Gonzalez", "maria.g@example.com", "(555) 203-8876", "2024-01-15", "$79.00", "Active", "97203", "Upgraded to annual"],
        ["M-1004", "James", "Patel", "james.patel@example.com", "(555) 204-4421", "2024-02-02", "$49.00", "Active", "97204", ""],
        ["M-1005", "Emily", "Nguyen", "emily.n@example.com", "(555) 205-9910", "2024-02-09", "$49.00", "Active", "97205", "Referred by M-1001"],
        # A few mixed date formats
        ["M-1006", "Robert", "Kim", "robert.kim@example.com", "(555) 206-3312", "02/14/2024", "$79.00", "Active", "97206", ""],
        ["M-1007", "Lisa", "Thompson", "lisa.t@example.com", "(555) 207-7788", "Feb 20, 2024", "$49.00", "Active", "97207", ""],
        ["M-1008", "Michael", "Brooks", "m.brooks@example.com", "(555) 208-1123", "03/01/2024", "$49.00", "Active", "97208", "Corporate plan"],
        # A few mixed phone formats
        ["M-1009", "Jennifer", "Diaz", "jen.diaz@example.com", "555-209-4456", "2024-03-05", "$49.00", "Active", "97209", ""],
        ["M-1010", "William", "Foster", "will.foster@example.com", "5552104078", "2024-03-11", "$79.00", "Active", "97210", ""],
        ["M-1011", "Ashley", "Rivera", "ashley.r@example.com", "555.211.6690", "2024-03-18", "$49.00", "Active", "97211", ""],
        # A couple mixed currency formats (no $, no decimals)
        ["M-1012", "Daniel", "Wright", "dan.wright@example.com", "(555) 212-3341", "2024-03-22", "49", "Active", "97212", ""],
        ["M-1013", "Nicole", "Adams", "nicole.a@example.com", "(555) 213-9982", "2024-04-01", "$79", "Active", "97213", ""],
        # Two placeholder cells (not a flood)
        ["M-1014", "Kevin", "Murphy", "kevin.m@example.com", "(555) 214-5567", "2024-04-08", "$49.00", "Active", "97214", "N/A"],
        ["M-1015", "Rachel", "Bennett", "rachel.b@example.com", "(555) 215-2203", "2024-04-15", "$49.00", "Active", "97215", "TBD"],
        # One lowercase status among "Active"
        ["M-1016", "Brandon", "Cole", "brandon.c@example.com", "(555) 216-8834", "2024-04-20", "$49.00", "active", "97216", ""],
        # Legitimate inactive members
        ["M-1017", "Megan", "Torres", "megan.t@example.com", "(555) 217-1145", "2024-05-02", "$49.00", "Inactive", "97217", "Cancelled"],
        ["M-1018", "Justin", "Reed", "justin.r@example.com", "(555) 218-6620", "2024-05-09", "$49.00", "Inactive", "97218", ""],
        # A little missing data (empty phone, empty zip)
        ["M-1019", "Amanda", "Price", "amanda.p@example.com", "", "2024-05-14", "$49.00", "Active", "97219", ""],
        ["M-1020", "Tyler", "Hughes", "tyler.h@example.com", "(555) 220-3398", "2024-05-21", "$49.00", "Active", "", ""],
        # More clean rows
        ["M-1021", "Olivia", "Ward", "olivia.w@example.com", "(555) 221-7712", "2024-06-03", "$79.00", "Active", "97221", "Annual"],
        ["M-1022", "Nathan", "Perry", "nathan.p@example.com", "(555) 222-4409", "2024-06-10", "$49.00", "Active", "97222", ""],
        ["M-1023", "Hannah", "Long", "hannah.l@example.com", "(555) 223-9987", "2024-06-17", "$49.00", "Active", "97223", ""],
        # One exact double-entry (identical to M-1001)
        ["M-1001", "Sarah", "Chen", "sarah.chen@example.com", "(555) 201-3345", "2024-01-08", "$49.00", "Active", "97201", "Founding member"],
        # One case-variant duplicate of M-1002 (David Okafor)
        ["M-1024", "david", "okafor", "D.OKAFOR@example.com", "(555) 202-1190", "2024-01-12", "$49.00", "Active", "97202", ""],
        ["M-1025", "Christina", "Bell", "christina.b@example.com", "(555) 225-1160", "2024-06-24", "$49.00", "Active", "97225", ""],
        ["M-1026", "Aaron", "Cooper", "aaron.c@example.com", "(555) 226-8845", "2024-07-01", "$49.00", "Active", "97226", ""],
        ["M-1027", "Victoria", "Gray", "victoria.g@example.com", "(555) 227-3390", "2024-07-08", "$79.00", "Active", "97227", ""],
    ]
    for row in rows:
        ws.append(row)

    return wb


def main():
    input_dir = Path(__file__).parent / "samples" / "input"
    input_dir.mkdir(parents=True, exist_ok=True)

    messy_path = input_dir / "sample_messy_data.xlsx"
    build_workbook().save(messy_path)
    print(f"Sample file generated: {messy_path}")

    realistic_path = input_dir / "sample_realistic_data.xlsx"
    build_realistic_workbook().save(realistic_path)
    print(f"Sample file generated: {realistic_path}")


if __name__ == "__main__":
    main()
