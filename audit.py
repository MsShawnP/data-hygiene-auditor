#!/usr/bin/env python3
"""
Data Hygiene Auditor — Lailara LLC
Detects mixed formats, misused fields, placeholder floods, and phantom duplicates
in Excel files. Outputs HTML report, Excel findings, and PDF.

Usage: python audit.py --input myfile.xlsx --output ./reports
"""

import argparse
import os
import re
import sys
import json
import hashlib
from collections import Counter, defaultdict
from datetime import datetime
from html import escape as _html_escape
from pathlib import Path
from xml.sax.saxutils import escape as _xml_escape

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors as rl_colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
)
from reportlab.lib.enums import TA_CENTER


def _h(val):
    """Escape a value for safe inclusion in HTML text or attributes."""
    return _html_escape(str(val), quote=True)


def _p(val):
    """Escape a value for inclusion inside a reportlab Paragraph."""
    return _xml_escape(str(val))


def _supports_color():
    """Check if the terminal supports ANSI color codes."""
    if os.environ.get('NO_COLOR'):
        return False
    if sys.platform == 'win32':
        os.system('')  # enables ANSI on Windows 10+
    return hasattr(sys.stdout, 'isatty') and sys.stdout.isatty()


_COLOR = _supports_color()


def _c(text, code):
    """Wrap text in ANSI color if supported."""
    if not _COLOR:
        return text
    return f"\033[{code}m{text}\033[0m"

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

PLACEHOLDER_PATTERNS = [
    r'^test$', r'^n/?a$', r'^null$', r'^none$', r'^tbd$', r'^xxx+$',
    r'^placeholder$', r'^default$', r'^dummy$', r'^sample$', r'^todo$',
    r'^temp$', r'^unknown$', r'^undefined$', r'^blank$', r'^\.+$',
    r'^0{3,}$', r'^-+$', r'^na$',
]
PLACEHOLDER_RE = re.compile('|'.join(PLACEHOLDER_PATTERNS), re.IGNORECASE)

DATE_PATTERNS = [
    (r'^\d{4}-\d{2}-\d{2}$', 'YYYY-MM-DD'),
    (r'^\d{2}/\d{2}/\d{4}$', 'MM/DD/YYYY'),
    (r'^\d{1,2}/\d{1,2}/\d{4}$', 'M/D/YYYY'),
    (r'^[A-Za-z]{3,9}\s+\d{1,2},?\s+\d{4}$', 'Mon DD, YYYY'),
    (r'^\d{1,2}-[A-Za-z]{3}-\d{4}$', 'DD-Mon-YYYY'),
    (r'^\d{4}/\d{2}/\d{2}$', 'YYYY/MM/DD'),
    (r'^\d{4}\.\d{2}\.\d{2}$', 'YYYY.MM.DD'),
    (r'^\d{2}-\d{2}-\d{4}$', 'MM-DD-YYYY'),
    (r'^[A-Za-z]{3,9}\s+\d{1,2}\s+\d{4}$', 'Mon DD YYYY'),
]

PHONE_PATTERNS = [
    (r'^\(\d{3}\)\s?\d{3}-\d{4}$', '(XXX) XXX-XXXX'),
    (r'^\d{3}-\d{3}-\d{4}$', 'XXX-XXX-XXXX'),
    (r'^\d{10}$', 'XXXXXXXXXX'),
    (r'^\d{3}\.\d{3}\.\d{4}$', 'XXX.XXX.XXXX'),
    (r'^\d{3}\s\d{3}\s\d{4}$', 'XXX XXX XXXX'),
    (r'^\+?1-?\d{3}-\d{3}-\d{4}$', '+1-XXX-XXX-XXXX'),
    (r'^\(\d{3}\)\d{7}$', '(XXX)XXXXXXX'),
]

CURRENCY_PATTERNS = [
    (r'^\$[\d,]+\.\d{2}$', '$X,XXX.XX'),
    (r'^\$[\d,]+$', '$X,XXX'),
    (r'^[\d,]+\.\d{2}$', 'X,XXX.XX (no symbol)'),
    (r'^[\d,]+$', 'X,XXX (bare number)'),
    (r'^\$-?[\d,]+\.\d{2}\s*(USD)?$', '$X,XXX.XX USD'),
    (r'^\$[\d,]+\.\d{2}\s+USD$', '$X,XXX.XX USD'),
]

EMAIL_RE = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')


# ─────────────────────────────────────────────────────────────────────────────
# FIELD TYPE INFERENCE
# ─────────────────────────────────────────────────────────────────────────────

def infer_field_type(col_name, values):
    """Infer the semantic type of a field from its name and sample values."""
    name_lower = col_name.lower().replace('_', '').replace(' ', '')
    non_null = [str(v).strip() for v in values if pd.notna(v) and str(v).strip()]
    if not non_null:
        return 'empty'
    sample = non_null[:200]

    # Name-based hints
    if any(k in name_lower for k in ['date', 'time', 'created', 'updated', 'ship', 'join', 'birth', 'dob']):
        return 'date'
    if any(k in name_lower for k in ['phone', 'fax', 'mobile', 'cell', 'tel']):
        return 'phone'
    if any(k in name_lower for k in ['email', 'mail']):
        return 'email'
    if any(k in name_lower for k in ['price', 'amount', 'balance', 'cost', 'fee', 'total', 'revenue', 'salary']):
        return 'currency'
    if any(k in name_lower for k in ['zip', 'postal']):
        return 'zipcode'
    if any(k in name_lower for k in ['id', 'code', 'key', 'sku', 'ref', 'number']) and 'phone' not in name_lower:
        return 'id'
    if any(k in name_lower for k in ['first', 'last', 'name', 'fname', 'lname']):
        return 'name'
    if any(k in name_lower for k in ['status', 'state', 'type', 'category', 'flag']):
        return 'categorical'
    if any(k in name_lower for k in ['note', 'comment', 'desc', 'remark', 'memo']):
        return 'freetext'

    # Content-based fallback
    date_hits = sum(1 for v in sample if any(re.match(p, v) for p, _ in DATE_PATTERNS))
    if date_hits > len(sample) * 0.5:
        return 'date'
    phone_hits = sum(1 for v in sample if any(re.match(p, v) for p, _ in PHONE_PATTERNS))
    if phone_hits > len(sample) * 0.5:
        return 'phone'
    curr_hits = sum(1 for v in sample if any(re.match(p, v) for p, _ in CURRENCY_PATTERNS))
    if curr_hits > len(sample) * 0.5:
        return 'currency'
    email_hits = sum(1 for v in sample if EMAIL_RE.match(v))
    if email_hits > len(sample) * 0.5:
        return 'email'

    return 'freetext'


# ─────────────────────────────────────────────────────────────────────────────
# ANALYSIS ENGINES
# ─────────────────────────────────────────────────────────────────────────────

def analyze_nulls(series):
    """Baseline null/missing analysis for every field."""
    total = len(series)
    null_count = series.isna().sum()
    whitespace_only = sum(1 for v in series if pd.notna(v) and len(str(v)) > 0 and str(v).strip() == '')
    blank_count = sum(1 for v in series if pd.notna(v) and str(v) == '')
    missing = null_count + blank_count + whitespace_only
    pct = (missing / total * 100) if total > 0 else 0
    return {
        'null_count': int(null_count),
        'blank_count': int(blank_count),
        'whitespace_only': int(whitespace_only),
        'total_missing': int(missing),
        'missing_pct': round(pct, 1),
        'total_rows': total,
    }


def analyze_mixed_formats(series, field_type):
    """Detect inconsistent formatting within a field."""
    non_null = [(i, str(v).strip()) for i, v in series.items() if pd.notna(v) and str(v).strip()]
    if not non_null:
        return None

    if field_type == 'date':
        patterns = DATE_PATTERNS
    elif field_type == 'phone':
        patterns = PHONE_PATTERNS
    elif field_type == 'currency':
        patterns = CURRENCY_PATTERNS
    else:
        return None

    format_counts = Counter()
    unmatched = []
    for idx, val in non_null:
        matched = False
        for regex, label in patterns:
            if re.match(regex, val):
                format_counts[label] += 1
                matched = True
                break
        if not matched:
            format_counts['(non-standard)'] += 1
            unmatched.append(val)

    if len(format_counts) <= 1:
        return None

    dominant = format_counts.most_common(1)[0]
    total_typed = sum(format_counts.values())
    inconsistent_count = total_typed - dominant[1]
    inconsistent_pct = round(inconsistent_count / total_typed * 100, 1)

    return {
        'field_type': field_type,
        'format_distribution': dict(format_counts.most_common()),
        'dominant_format': dominant[0],
        'dominant_count': dominant[1],
        'inconsistent_count': inconsistent_count,
        'inconsistent_pct': inconsistent_pct,
        'sample_nonstandard': unmatched[:5],
    }


def analyze_wrong_purpose(series, col_name, field_type):
    """Detect fields being used for the wrong purpose."""
    findings = []
    non_null = [(i, str(v).strip()) for i, v in series.items() if pd.notna(v) and str(v).strip()]
    if not non_null:
        return findings

    if field_type == 'name':
        for idx, val in non_null:
            if re.match(r'^[A-Z]{2,}-\d+', val) or re.match(r'^REF-', val, re.IGNORECASE):
                findings.append({
                    'issue': 'Code/ID stuffed in name field',
                    'example': val,
                    'row': idx,
                })
            elif re.match(r'^\d+$', val):
                findings.append({
                    'issue': 'Numeric value in name field',
                    'example': val,
                    'row': idx,
                })

    if field_type == 'currency':
        for idx, val in non_null:
            if re.match(r'^[a-zA-Z]', val):
                findings.append({
                    'issue': 'Text in currency field',
                    'example': val,
                    'row': idx,
                })

    if field_type == 'email':
        for idx, val in non_null:
            if not EMAIL_RE.match(val) and not PLACEHOLDER_RE.match(val):
                findings.append({
                    'issue': 'Invalid email format',
                    'example': val,
                    'row': idx,
                })

    if field_type == 'id':
        type_counts = Counter()
        for idx, val in non_null:
            if re.match(r'^[A-Za-z]+-\d+$', val):
                type_counts['alphanumeric_code'] += 1
            elif re.match(r'^\d+$', val):
                type_counts['bare_number'] += 1
            else:
                type_counts['other'] += 1
        if len(type_counts) > 1:
            label_map = {
                'alphanumeric_code': 'coded (e.g. CUST-001)',
                'bare_number': 'bare numbers',
                'other': 'other',
            }
            parts = [f"{type_counts[k]} {label_map[k]}" for k, _ in type_counts.most_common()]
            findings.append({
                'issue': 'Mixed ID formats',
                'example': ' vs '.join(parts),
                'row': None,
            })

    if field_type == 'categorical':
        unique_vals = set(v.lower() for _, v in non_null)
        if unique_vals and all(v in ('0', '1', 'y', 'n', 'yes', 'no', 'true', 'false') for v in unique_vals):
            raw_vals = set(v for _, v in non_null)
            bool_types = set()
            for v in raw_vals:
                vl = v.lower()
                if vl in ('0', '1'): bool_types.add('numeric')
                elif vl in ('y', 'n'): bool_types.add('y/n')
                elif vl in ('yes', 'no'): bool_types.add('yes/no')
                elif vl in ('true', 'false'): bool_types.add('true/false')
            if len(bool_types) > 1:
                findings.append({
                    'issue': 'Mixed boolean representations',
                    'example': ', '.join(raw_vals),
                    'row': None,
                })
        case_variants = defaultdict(set)
        for _, v in non_null:
            case_variants[v.lower()].add(v)
        for key, variants in case_variants.items():
            if len(variants) > 1:
                findings.append({
                    'issue': 'Inconsistent casing in categorical values',
                    'example': ' / '.join(sorted(variants)),
                    'row': None,
                })

    return findings


def analyze_placeholders(series, col_name):
    """Detect suspiciously uniform or placeholder data."""
    findings = []
    non_null = [str(v).strip() for v in series if pd.notna(v) and str(v).strip()]
    if not non_null:
        return findings

    placeholder_hits = [v for v in non_null if PLACEHOLDER_RE.match(v)]
    if placeholder_hits:
        counter = Counter(placeholder_hits)
        for val, count in counter.most_common(5):
            findings.append({
                'type': 'placeholder_value',
                'value': val,
                'count': count,
                'pct': round(count / len(non_null) * 100, 1),
            })

    val_counts = Counter(non_null)
    total = len(non_null)
    for val, count in val_counts.most_common(5):
        if count >= 3 and (count / total) >= 0.1 and not PLACEHOLDER_RE.match(val):
            findings.append({
                'type': 'suspicious_repetition',
                'value': val,
                'count': count,
                'pct': round(count / total * 100, 1),
            })

    return findings


def analyze_phantom_duplicates(df, sheet_name, field_types=None):
    """Detect records that are the same after normalizing whitespace/case/punctuation.
    Excludes ID/key columns from matching so records with different IDs but identical
    content are still caught."""
    findings = []
    if df.empty or len(df) < 2:
        return findings

    field_types = field_types or {}

    def normalize(val):
        if pd.isna(val):
            return ''
        s = str(val).strip().lower()
        s = re.sub(r'\s+', ' ', s)
        s = re.sub(r'[^\w\s@.]', '', s)
        return s

    # Exclude ID-type columns from signature — they're expected to be unique
    id_cols = set()
    for col in df.columns:
        ft = field_types.get(col, infer_field_type(col, df[col].values))
        if ft == 'id':
            id_cols.add(col)
        # Also exclude columns that are completely unique (likely surrogate keys)
        elif df[col].nunique() == len(df):
            id_cols.add(col)

    content_cols = [c for c in df.columns if c not in id_cols]
    if not content_cols:
        content_cols = list(df.columns)

    normalized = df[content_cols].apply(lambda col: col.map(normalize))

    # Drop columns where every value is unique (no chance of matching)
    sig_cols = [c for c in content_cols if normalized[c].nunique() < len(df)]
    if not sig_cols:
        return findings

    norm_subset = normalized[sig_cols]
    sigs = norm_subset.apply(lambda row: hashlib.md5('||'.join(row.values).encode()).hexdigest(), axis=1)
    dup_sigs = sigs[sigs.duplicated(keep=False)]
    if dup_sigs.empty:
        return findings

    groups = defaultdict(list)
    for idx, sig in dup_sigs.items():
        groups[sig].append(idx)

    for sig, indices in groups.items():
        if len(indices) < 2:
            continue
        row_nums = [i + 2 for i in indices]  # +2 for header + 0-index
        sample_rows = []
        for i in indices[:3]:
            row_data = {col: str(df.iloc[i][col]) for col in df.columns[:6]}
            sample_rows.append(row_data)

        # Check if raw (unnormalized) content columns match exactly
        raw_match = True
        first_raw = tuple(str(df.iloc[indices[0]][c]) for c in content_cols)
        for i in indices[1:]:
            if tuple(str(df.iloc[i][c]) for c in content_cols) != first_raw:
                raw_match = False
                break

        findings.append({
            'rows': row_nums,
            'group_size': len(indices),
            'exact_match': raw_match,
            'sample_data': sample_rows,
            'matched_on': sig_cols,
            'excluded_id_cols': list(id_cols),
            'type': 'exact_duplicate' if raw_match else 'phantom_duplicate',
        })

    return findings


# ─────────────────────────────────────────────────────────────────────────────
# SEVERITY RATING
# ─────────────────────────────────────────────────────────────────────────────

def rate_severity(finding_type, details):
    """Assign High / Medium / Low severity."""
    if finding_type == 'phantom_duplicate':
        if details.get('type') == 'exact_duplicate':
            return 'High'
        return 'High' if details.get('group_size', 0) > 3 else 'Medium'

    if finding_type == 'mixed_format':
        pct = details.get('inconsistent_pct', 0)
        if pct > 30:
            return 'High'
        elif pct > 10:
            return 'Medium'
        return 'Low'

    if finding_type == 'wrong_purpose':
        return 'High'

    if finding_type == 'placeholder':
        pct = details.get('pct', 0)
        if pct > 20:
            return 'High'
        elif pct > 5:
            return 'Medium'
        return 'Low'

    if finding_type == 'null_analysis':
        pct = details.get('missing_pct', 0)
        if pct > 50:
            return 'High'
        elif pct > 20:
            return 'Medium'
        elif pct > 5:
            return 'Low'
        return None  # Don't flag low-missing fields

    return 'Medium'


# ─────────────────────────────────────────────────────────────────────────────
# WHY-IT-MATTERS EXPLANATIONS
# ─────────────────────────────────────────────────────────────────────────────

WHY_IT_MATTERS = {
    'mixed_format_date': (
        "Mixed date formats cause sorting failures, broken filters, and incorrect calculations. "
        "A date stored as text (\"Jan 15, 2023\") won't sort chronologically next to \"2023-01-15\". "
        "Downstream tools, APIs, and reports will misparse or reject inconsistent dates."
    ),
    'mixed_format_phone': (
        "Inconsistent phone formats break deduplication, prevent reliable search/lookup, and cause "
        "issues with automated dialers or SMS systems. Two records for the same person may appear "
        "as different contacts if one says \"(555) 123-4567\" and another says \"5551234567\"."
    ),
    'mixed_format_currency': (
        "Mixed currency formats (\"$1,250.00\" vs \"1250\" vs \"five thousand\") prevent accurate "
        "aggregation and comparison. Summing a column with text-formatted currency returns errors or "
        "silently drops values, leading to wrong totals in financial reports."
    ),
    'wrong_purpose': (
        "When a field is used for something other than its intended purpose — like storing reference "
        "codes in a name field or text in a currency field — it corrupts both the misused field and "
        "whatever field should have held that data. This makes the data unreliable for any analysis."
    ),
    'placeholder': (
        "Placeholder values (\"Test\", \"N/A\", \"TBD\") that persist in production data inflate counts, "
        "skew averages, and create phantom records. They often indicate incomplete data entry or "
        "inadequate validation at the point of capture."
    ),
    'suspicious_repetition': (
        "When the same value appears far more often than expected, it may indicate a default value "
        "that was never updated, a copy-paste error, or a system glitch that stamped the same data "
        "across multiple records."
    ),
    'phantom_duplicate': (
        "These records look different on the surface (different casing, extra spaces, punctuation "
        "variations) but represent the same entity. They cause inflated counts, split transaction "
        "histories, and duplicate outreach — problems that compound over time."
    ),
    'exact_duplicate': (
        "Exact duplicate rows are the clearest sign of a data quality issue — they can result from "
        "double-submissions, ETL failures, or missing unique constraints. Every duplicate inflates "
        "counts and distorts any metric built on this data."
    ),
    'null_analysis': (
        "High rates of missing data reduce the reliability of any analysis built on this field. "
        "Missing values can skew averages, break joins between tables, and cause downstream systems "
        "to error out or produce incomplete results."
    ),
}


# ─────────────────────────────────────────────────────────────────────────────
# MAIN AUDIT ORCHESTRATOR
# ─────────────────────────────────────────────────────────────────────────────

SUPPORTED_EXTENSIONS = {'.xlsx', '.xls', '.csv', '.tsv'}


def _load_sheets(input_path):
    """Load tabular data as a dict of {sheet_name: DataFrame}."""
    ext = Path(input_path).suffix.lower()
    if ext in ('.csv', '.tsv'):
        sep = '\t' if ext == '.tsv' else ','
        df = pd.read_csv(input_path, dtype=str, sep=sep)
        return {Path(input_path).stem: df}
    else:
        xls = pd.ExcelFile(input_path)
        return {name: pd.read_excel(xls, sheet_name=name, dtype=str)
                for name in xls.sheet_names}


def run_audit(input_path):
    """Run all checks against an Excel or CSV file. Returns structured audit results."""
    sheets = _load_sheets(input_path)
    results = {
        'input_file': os.path.basename(input_path),
        'audit_timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'sheets': {},
    }

    for sheet_name, df in sheets.items():
        if df.empty:
            continue

        sheet_results = {
            'row_count': len(df),
            'col_count': len(df.columns),
            'fields': {},
            'phantom_duplicates': [],
        }

        for col in df.columns:
            field_type = infer_field_type(col, df[col].values)
            null_info = analyze_nulls(df[col])
            mixed = analyze_mixed_formats(df[col], field_type)
            wrong = analyze_wrong_purpose(df[col], col, field_type)
            placeholders = analyze_placeholders(df[col], col)

            field_findings = {
                'inferred_type': field_type,
                'null_analysis': null_info,
                'issues': [],
            }

            null_severity = rate_severity('null_analysis', null_info)
            if null_severity:
                field_findings['issues'].append({
                    'type': 'null_analysis',
                    'severity': null_severity,
                    'detail': null_info,
                    'why': WHY_IT_MATTERS['null_analysis'],
                })

            if mixed:
                sev = rate_severity('mixed_format', mixed)
                why_key = f'mixed_format_{field_type}'
                field_findings['issues'].append({
                    'type': 'mixed_format',
                    'severity': sev,
                    'detail': mixed,
                    'why': WHY_IT_MATTERS.get(why_key, f'Mixed {field_type} formats reduce data consistency and can cause errors in downstream processing.'),
                })

            for w in wrong:
                field_findings['issues'].append({
                    'type': 'wrong_purpose',
                    'severity': rate_severity('wrong_purpose', w),
                    'detail': w,
                    'why': WHY_IT_MATTERS['wrong_purpose'],
                })

            for p in placeholders:
                ptype = p.get('type', 'placeholder')
                sev = rate_severity('placeholder', p)
                why_key = 'suspicious_repetition' if ptype == 'suspicious_repetition' else 'placeholder'
                field_findings['issues'].append({
                    'type': ptype,
                    'severity': sev,
                    'detail': p,
                    'why': WHY_IT_MATTERS[why_key],
                })

            sheet_results['fields'][col] = field_findings

        # Build field type map for dedup
        field_types = {col: fd['inferred_type'] for col, fd in sheet_results['fields'].items()}
        dupes = analyze_phantom_duplicates(df, sheet_name, field_types)
        for d in dupes:
            d['severity'] = rate_severity('phantom_duplicate', d)
            d['why'] = WHY_IT_MATTERS.get(d['type'], WHY_IT_MATTERS['phantom_duplicate'])
        sheet_results['phantom_duplicates'] = dupes

        results['sheets'][sheet_name] = sheet_results

    return results


# ─────────────────────────────────────────────────────────────────────────────
# HTML REPORT GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

def generate_html(results, output_path):
    """Generate a client-readable HTML report."""
    total_issues = 0
    severity_totals = Counter()
    for sheet in results['sheets'].values():
        for field in sheet['fields'].values():
            for issue in field['issues']:
                total_issues += 1
                severity_totals[issue['severity']] += 1
        for d in sheet['phantom_duplicates']:
            total_issues += 1
            severity_totals[d['severity']] += 1

    parts = []
    parts.append(f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Data Hygiene Audit — {_h(results['input_file'])}</title>
<style>
:root {{
    --bg: #1a1a2e;
    --card: #16213e;
    --card-border: #0f3460;
    --text: #e0e0e0;
    --text-muted: #8892a0;
    --accent: #e94560;
    --accent-warm: #d4a574;
    --high: #DC3545;
    --medium: #FFC107;
    --low: #28A745;
    --info: #4a90d9;
}}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    background: var(--bg);
    color: var(--text);
    line-height: 1.6;
    padding: 2rem;
    max-width: 1200px;
    margin: 0 auto;
}}
h1 {{ color: var(--accent); font-size: 1.8rem; margin-bottom: 0.25rem; }}
h2 {{ color: var(--accent-warm); font-size: 1.4rem; margin: 2rem 0 1rem; border-bottom: 1px solid var(--card-border); padding-bottom: 0.5rem; }}
h3 {{ color: var(--text); font-size: 1.1rem; margin: 1.5rem 0 0.5rem; }}
.subtitle {{ color: var(--text-muted); font-size: 0.95rem; margin-bottom: 1.5rem; }}
.summary-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: 1rem;
    margin: 1.5rem 0;
}}
.summary-card {{
    background: var(--card);
    border: 1px solid var(--card-border);
    border-radius: 8px;
    padding: 1.2rem;
    text-align: center;
}}
.summary-card .number {{ font-size: 2rem; font-weight: 700; }}
.summary-card .label {{ color: var(--text-muted); font-size: 0.85rem; text-transform: uppercase; letter-spacing: 0.05em; }}
.high .number {{ color: var(--high); }}
.medium .number {{ color: var(--medium); }}
.low .number {{ color: var(--low); }}
.info .number {{ color: var(--info); }}
.field-card {{
    background: var(--card);
    border: 1px solid var(--card-border);
    border-radius: 8px;
    padding: 1.2rem;
    margin-bottom: 1rem;
}}
.field-header {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 0.75rem;
}}
.field-name {{ font-weight: 600; font-size: 1.05rem; }}
.field-type {{
    background: var(--card-border);
    color: var(--text-muted);
    padding: 0.2rem 0.6rem;
    border-radius: 12px;
    font-size: 0.8rem;
}}
.null-bar {{
    height: 6px;
    background: #2a2a4a;
    border-radius: 3px;
    margin: 0.5rem 0;
    overflow: hidden;
}}
.null-bar-fill {{
    height: 100%;
    border-radius: 3px;
    transition: width 0.3s;
}}
.issue {{
    border-left: 3px solid var(--text-muted);
    padding: 0.75rem 1rem;
    margin: 0.75rem 0;
    background: rgba(255,255,255,0.02);
    border-radius: 0 6px 6px 0;
}}
.issue.severity-High {{ border-left-color: var(--high); }}
.issue.severity-Medium {{ border-left-color: var(--medium); }}
.issue.severity-Low {{ border-left-color: var(--low); }}
.severity-badge {{
    display: inline-block;
    padding: 0.15rem 0.5rem;
    border-radius: 10px;
    font-size: 0.75rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.05em;
}}
.severity-badge.High {{ background: var(--high); color: #fff; }}
.severity-badge.Medium {{ background: var(--medium); color: #000; }}
.severity-badge.Low {{ background: var(--low); color: #fff; }}
.why-box {{
    margin-top: 0.5rem;
    padding: 0.5rem 0.75rem;
    background: rgba(74, 144, 217, 0.08);
    border-radius: 4px;
    font-size: 0.9rem;
    color: var(--text-muted);
}}
.why-box strong {{ color: var(--info); }}
.format-table {{
    width: 100%;
    border-collapse: collapse;
    margin: 0.5rem 0;
    font-size: 0.9rem;
}}
.format-table th, .format-table td {{
    text-align: left;
    padding: 0.4rem 0.75rem;
    border-bottom: 1px solid var(--card-border);
}}
.format-table th {{ color: var(--text-muted); font-weight: 600; }}
.dup-group {{
    background: var(--card);
    border: 1px solid var(--card-border);
    border-radius: 8px;
    padding: 1rem;
    margin-bottom: 1rem;
}}
.footer {{
    margin-top: 3rem;
    padding-top: 1rem;
    border-top: 1px solid var(--card-border);
    color: var(--text-muted);
    font-size: 0.85rem;
    text-align: center;
}}
</style>
</head>
<body>

<h1>Data Hygiene Audit Report</h1>
<p class="subtitle">{_h(results['input_file'])} &mdash; {results['audit_timestamp']}</p>

<div class="summary-grid">
    <div class="summary-card info"><div class="number">{total_issues}</div><div class="label">Total Issues</div></div>
    <div class="summary-card high"><div class="number">{severity_totals.get('High', 0)}</div><div class="label">High Severity</div></div>
    <div class="summary-card medium"><div class="number">{severity_totals.get('Medium', 0)}</div><div class="label">Medium Severity</div></div>
    <div class="summary-card low"><div class="number">{severity_totals.get('Low', 0)}</div><div class="label">Low Severity</div></div>
</div>
""")

    for sheet_name, sheet_data in results['sheets'].items():
        parts.append(f"""
<h2>Sheet: {_h(sheet_name)}</h2>
<p style="color:var(--text-muted);margin-bottom:1rem;">{sheet_data['row_count']} rows &times; {sheet_data['col_count']} columns</p>
""")
        for col_name, field_data in sheet_data['fields'].items():
            null = field_data['null_analysis']
            issues = field_data['issues']
            ftype = field_data['inferred_type']

            null_color = 'var(--low)' if null['missing_pct'] < 10 else ('var(--medium)' if null['missing_pct'] < 30 else 'var(--high)')

            parts.append(f"""
<div class="field-card">
    <div class="field-header">
        <span class="field-name">{_h(col_name)}</span>
        <span class="field-type">{_h(ftype)}</span>
    </div>
    <div style="font-size:0.85rem;color:var(--text-muted);">
        Missing: {null['total_missing']} / {null['total_rows']} ({null['missing_pct']}%)
        {f" &mdash; {null['whitespace_only']} whitespace-only" if null['whitespace_only'] else ""}
    </div>
    <div class="null-bar"><div class="null-bar-fill" style="width:{min(null['missing_pct'], 100)}%;background:{null_color};"></div></div>
""")
            for issue in issues:
                sev = issue['severity']
                itype = issue['type']
                detail = issue['detail']
                why = issue.get('why', '')

                parts.append(f'<div class="issue severity-{sev}">')
                parts.append(f'<span class="severity-badge {sev}">{sev}</span> ')

                if itype == 'mixed_format':
                    parts.append(f'<strong>Mixed {_h(detail["field_type"])} formats</strong> &mdash; {detail["inconsistent_count"]} of {detail["dominant_count"] + detail["inconsistent_count"]} values deviate from dominant format ({_h(detail["dominant_format"])})')
                    parts.append('<table class="format-table"><tr><th>Format</th><th>Count</th></tr>')
                    for fmt, cnt in detail['format_distribution'].items():
                        parts.append(f'<tr><td>{_h(fmt)}</td><td>{cnt}</td></tr>')
                    parts.append('</table>')
                    if detail.get('sample_nonstandard'):
                        samples = ", ".join(_h(s) for s in detail["sample_nonstandard"][:3])
                        parts.append(f'<div style="font-size:0.85rem;color:var(--text-muted);">Non-standard samples: {samples}</div>')

                elif itype == 'wrong_purpose':
                    parts.append(f'<strong>{_h(detail["issue"])}</strong>')
                    if detail.get('example'):
                        parts.append(f' &mdash; e.g. "{_h(detail["example"])}"')
                    if detail.get('row') is not None:
                        parts.append(f' (row {detail["row"] + 2})')

                elif itype in ('placeholder_value', 'placeholder'):
                    parts.append(f'<strong>Placeholder detected:</strong> "{_h(detail["value"])}" appears {detail["count"]} times ({detail["pct"]}%)')

                elif itype == 'suspicious_repetition':
                    parts.append(f'<strong>Suspicious repetition:</strong> "{_h(detail["value"])}" appears {detail["count"]} times ({detail["pct"]}%)')

                elif itype == 'null_analysis':
                    parts.append(f'<strong>High missing rate:</strong> {detail["total_missing"]} of {detail["total_rows"]} values missing ({detail["missing_pct"]}%)')

                else:
                    parts.append(f'<strong>{_h(itype)}</strong>: {_h(json.dumps(detail, default=str))}')

                if why:
                    parts.append(f'<div class="why-box"><strong>Why this matters:</strong> {_h(why)}</div>')
                parts.append('</div>')

            parts.append('</div>')

        if sheet_data['phantom_duplicates']:
            parts.append('<h3>Phantom &amp; Exact Duplicates</h3>')
            for dup in sheet_data['phantom_duplicates']:
                sev = dup['severity']
                dtype = 'Exact Duplicate' if dup['type'] == 'exact_duplicate' else 'Phantom Duplicate'
                parts.append(f"""
<div class="dup-group">
    <span class="severity-badge {sev}">{sev}</span>
    <strong>{dtype}</strong> &mdash; {dup['group_size']} rows: {', '.join(str(r) for r in dup['rows'])}
    <table class="format-table">
        <tr>{''.join(f'<th>{_h(k)}</th>' for k in dup['sample_data'][0].keys())}</tr>
""")
                for row in dup['sample_data']:
                    parts.append('<tr>' + ''.join(f'<td>{_h(v)}</td>' for v in row.values()) + '</tr>')
                parts.append('</table>')
                parts.append(f'<div class="why-box"><strong>Why this matters:</strong> {_h(dup["why"])}</div>')
                parts.append('</div>')

    parts.append(f"""
<div class="footer">
    Data Hygiene Audit &mdash; Generated {results['audit_timestamp']} &mdash; Lailara LLC
</div>
</body></html>""")

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(''.join(parts))
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL FINDINGS GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel(results, output_path):
    """Generate sortable/filterable Excel findings file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"

    headers = ["Sheet", "Field", "Inferred Type", "Issue Type", "Severity",
               "Description", "Example / Detail", "Why It Matters"]
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="0f3460")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='cccccc'),
        right=Side(style='thin', color='cccccc'),
        top=Side(style='thin', color='cccccc'),
        bottom=Side(style='thin', color='cccccc'),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    sev_fills = {
        'High': PatternFill("solid", fgColor="F8D7DA"),
        'Medium': PatternFill("solid", fgColor="FFF3CD"),
        'Low': PatternFill("solid", fgColor="D4EDDA"),
    }

    row_num = 2
    for sheet_name, sheet_data in results['sheets'].items():
        for col_name, field_data in sheet_data['fields'].items():
            for issue in field_data['issues']:
                detail = issue['detail']
                itype = issue['type']

                if itype == 'mixed_format':
                    desc = f"Mixed {detail['field_type']} formats: {detail['inconsistent_count']} values deviate from {detail['dominant_format']}"
                    example = '; '.join(f"{k}: {v}" for k, v in detail['format_distribution'].items())
                elif itype == 'wrong_purpose':
                    desc = detail['issue']
                    example = detail.get('example', '')
                elif itype in ('placeholder_value', 'placeholder'):
                    desc = f"Placeholder \"{detail['value']}\" found {detail['count']} times"
                    example = f"{detail['pct']}% of non-null values"
                elif itype == 'suspicious_repetition':
                    desc = f"\"{detail['value']}\" repeated {detail['count']} times"
                    example = f"{detail['pct']}% of non-null values"
                elif itype == 'null_analysis':
                    desc = f"{detail['total_missing']} of {detail['total_rows']} values missing ({detail['missing_pct']}%)"
                    example = f"Null: {detail['null_count']}, Blank: {detail['blank_count']}, Whitespace: {detail['whitespace_only']}"
                else:
                    desc = str(itype)
                    example = json.dumps(detail, default=str)

                values = [sheet_name, col_name, field_data['inferred_type'],
                          itype, issue['severity'], desc, example, issue.get('why', '')]
                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    cell.font = Font(name="Arial", size=10)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = thin_border
                    if col_idx == 5:
                        cell.fill = sev_fills.get(issue['severity'], PatternFill())

                row_num += 1

        for dup in sheet_data['phantom_duplicates']:
            dtype = 'Exact Duplicate' if dup['type'] == 'exact_duplicate' else 'Phantom Duplicate'
            desc = f"{dup['group_size']} rows appear to be the same record"
            example = f"Rows: {', '.join(str(r) for r in dup['rows'])}"
            values = [sheet_name, "(row-level)", "—", dtype, dup['severity'],
                      desc, example, dup.get('why', '')]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
                if col_idx == 5:
                    cell.fill = sev_fills.get(dup['severity'], PatternFill())
            row_num += 1

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 35
    ws.column_dimensions['H'].width = 55

    ws.auto_filter.ref = f"A1:H{row_num - 1}"
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary", 0)
    ws2['A1'] = "Data Hygiene Audit Summary"
    ws2['A1'].font = Font(bold=True, size=14, name="Arial")
    ws2['A3'] = "File:"
    ws2['B3'] = results['input_file']
    ws2['A4'] = "Audit Date:"
    ws2['B4'] = results['audit_timestamp']
    ws2['A5'] = "Total Issues:"
    ws2['B5'] = row_num - 2
    for r in range(3, 6):
        ws2.cell(row=r, column=1).font = Font(bold=True, name="Arial", size=10)
        ws2.cell(row=r, column=2).font = Font(name="Arial", size=10)

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 40

    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# PDF REPORT GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

def generate_pdf(results, output_path):
    """Generate a clean PDF report matching the HTML content."""
    doc = SimpleDocTemplate(
        output_path, pagesize=letter,
        leftMargin=0.75*inch, rightMargin=0.75*inch,
        topMargin=0.75*inch, bottomMargin=0.75*inch,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='SectionHead', parent=styles['Heading2'],
                              textColor=rl_colors.HexColor('#d4a574'), fontSize=14,
                              spaceAfter=8, spaceBefore=16))
    styles.add(ParagraphStyle(name='FieldHead', parent=styles['Heading3'],
                              fontSize=11, spaceAfter=4, spaceBefore=10))
    styles.add(ParagraphStyle(name='SmallBody', parent=styles['Normal'],
                              fontSize=9, leading=12, spaceAfter=4))
    styles.add(ParagraphStyle(name='WhyBox', parent=styles['Normal'],
                              fontSize=8.5, leading=11, leftIndent=12,
                              textColor=rl_colors.HexColor('#555555'),
                              spaceAfter=6, spaceBefore=2))
    styles.add(ParagraphStyle(name='SevHigh', parent=styles['Normal'],
                              fontSize=9, textColor=rl_colors.HexColor('#DC3545'),
                              fontName='Helvetica-Bold'))
    styles.add(ParagraphStyle(name='SevMedium', parent=styles['Normal'],
                              fontSize=9, textColor=rl_colors.HexColor('#856404'),
                              fontName='Helvetica-Bold'))
    styles.add(ParagraphStyle(name='SevLow', parent=styles['Normal'],
                              fontSize=9, textColor=rl_colors.HexColor('#155724'),
                              fontName='Helvetica-Bold'))

    story = []

    story.append(Paragraph("Data Hygiene Audit Report", styles['Title']))
    story.append(Paragraph(f"{_p(results['input_file'])} — {results['audit_timestamp']}", styles['Normal']))
    story.append(Spacer(1, 12))

    total_issues = 0
    severity_totals = Counter()
    for sheet in results['sheets'].values():
        for field in sheet['fields'].values():
            for issue in field['issues']:
                total_issues += 1
                severity_totals[issue['severity']] += 1
        for d in sheet['phantom_duplicates']:
            total_issues += 1
            severity_totals[d['severity']] += 1

    summary_data = [
        ['Total Issues', 'High', 'Medium', 'Low'],
        [str(total_issues), str(severity_totals.get('High', 0)),
         str(severity_totals.get('Medium', 0)), str(severity_totals.get('Low', 0))],
    ]
    t = Table(summary_data, colWidths=[1.5*inch]*4)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#0f3460')),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#cccccc')),
        ('BACKGROUND', (1, 1), (1, 1), rl_colors.HexColor('#F8D7DA')),
        ('BACKGROUND', (2, 1), (2, 1), rl_colors.HexColor('#FFF3CD')),
        ('BACKGROUND', (3, 1), (3, 1), rl_colors.HexColor('#D4EDDA')),
    ]))
    story.append(t)
    story.append(Spacer(1, 16))

    for sheet_name, sheet_data in results['sheets'].items():
        story.append(Paragraph(f"Sheet: {_p(sheet_name)}", styles['SectionHead']))
        story.append(Paragraph(
            f"{sheet_data['row_count']} rows × {sheet_data['col_count']} columns",
            styles['SmallBody']))

        for col_name, field_data in sheet_data['fields'].items():
            issues = field_data['issues']
            if not issues:
                continue

            null = field_data['null_analysis']
            ftype = field_data['inferred_type']

            story.append(Paragraph(
                f"<b>{_p(col_name)}</b> <i>({_p(ftype)})</i> — Missing: {null['total_missing']}/{null['total_rows']} ({null['missing_pct']}%)",
                styles['FieldHead']))

            for issue in issues:
                sev = issue['severity']
                detail = issue['detail']
                itype = issue['type']
                sev_style = f'Sev{sev}'

                if itype == 'mixed_format':
                    text = f"[{sev}] Mixed {_p(detail['field_type'])} formats — {detail['inconsistent_count']} values deviate from {_p(detail['dominant_format'])}"
                    story.append(Paragraph(text, styles.get(sev_style, styles['SmallBody'])))
                    fmt_data = [['Format', 'Count']]
                    for fmt, cnt in detail['format_distribution'].items():
                        fmt_data.append([fmt, str(cnt)])
                    ft = Table(fmt_data, colWidths=[3*inch, 1*inch])
                    ft.setStyle(TableStyle([
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#e8e8e8')),
                        ('GRID', (0, 0), (-1, -1), 0.25, rl_colors.HexColor('#cccccc')),
                    ]))
                    story.append(ft)

                elif itype == 'wrong_purpose':
                    text = f"[{sev}] {_p(detail['issue'])}"
                    if detail.get('example'):
                        text += f' — e.g. "{_p(detail["example"])}"'
                    story.append(Paragraph(text, styles.get(sev_style, styles['SmallBody'])))

                elif itype in ('placeholder_value', 'placeholder'):
                    text = f'[{sev}] Placeholder: "{_p(detail["value"])}" × {detail["count"]} ({detail["pct"]}%)'
                    story.append(Paragraph(text, styles.get(sev_style, styles['SmallBody'])))

                elif itype == 'suspicious_repetition':
                    text = f'[{sev}] Repetition: "{_p(detail["value"])}" × {detail["count"]} ({detail["pct"]}%)'
                    story.append(Paragraph(text, styles.get(sev_style, styles['SmallBody'])))

                elif itype == 'null_analysis':
                    text = f"[{sev}] High missing rate: {detail['total_missing']}/{detail['total_rows']} ({detail['missing_pct']}%)"
                    story.append(Paragraph(text, styles.get(sev_style, styles['SmallBody'])))

                why = issue.get('why', '')
                if why:
                    story.append(Paragraph(f"<b>Why this matters:</b> {_p(why)}", styles['WhyBox']))

        if sheet_data['phantom_duplicates']:
            story.append(Paragraph("Phantom &amp; Exact Duplicates", styles['FieldHead']))
            for dup in sheet_data['phantom_duplicates']:
                sev = dup['severity']
                dtype = 'Exact Duplicate' if dup['type'] == 'exact_duplicate' else 'Phantom Duplicate'
                text = f"[{sev}] {dtype} — {dup['group_size']} rows: {', '.join(str(r) for r in dup['rows'])}"
                story.append(Paragraph(text, styles.get(f'Sev{sev}', styles['SmallBody'])))
                if dup.get('why'):
                    story.append(Paragraph(f"<b>Why this matters:</b> {_p(dup['why'])}", styles['WhyBox']))

        story.append(PageBreak())

    story.append(Paragraph(
        f"Data Hygiene Audit — Generated {results['audit_timestamp']} — Lailara LLC",
        ParagraphStyle(name='Footer', parent=styles['Normal'],
                       fontSize=8, textColor=rl_colors.HexColor('#999999'),
                       alignment=TA_CENTER)))

    doc.build(story)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# CLI ENTRYPOINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='Data Hygiene Auditor — Detect data quality issues in Excel and CSV files',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python audit.py --input customers.xlsx --output ./reports
  python audit.py --input data.csv --output ./reports
  python audit.py --input data.xlsx --output ./reports --json

Outputs three files:
  - audit_report.html   (visual, client-readable)
  - audit_findings.xlsx (sortable/filterable issue list)
  - audit_report.pdf    (email-ready deliverable)
        """
    )
    parser.add_argument('--input', '-i', required=True, help='Path to input file (.xlsx, .csv, .tsv)')
    parser.add_argument('--output', '-o', required=True, help='Output directory for reports')
    parser.add_argument('--json', action='store_true', help='Also output raw JSON results')
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"Error: Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    ext = Path(args.input).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        print(f"Error: Unsupported file type '{ext}'. Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}", file=sys.stderr)
        sys.exit(1)

    os.makedirs(args.output, exist_ok=True)

    basename = Path(args.input).stem
    print(f"\n  {_c('Data Hygiene Auditor', '1')}")
    print(f"  Auditing: {_c(args.input, '36')}\n")

    results = run_audit(args.input)
    sheet_count = len(results['sheets'])
    for i, name in enumerate(results['sheets'], 1):
        print(f"  [{i}/{sheet_count}] Analyzed sheet: {_c(name, '36')}")

    html_path = os.path.join(args.output, f"{basename}_audit_report.html")
    xlsx_path = os.path.join(args.output, f"{basename}_audit_findings.xlsx")
    pdf_path = os.path.join(args.output, f"{basename}_audit_report.pdf")

    print(f"\n  Generating reports...")

    generate_html(results, html_path)
    print(f"    {_c('HTML', '32')}  -> {html_path}")

    generate_excel(results, xlsx_path)
    print(f"    {_c('Excel', '32')} -> {xlsx_path}")

    generate_pdf(results, pdf_path)
    print(f"    {_c('PDF', '32')}   -> {pdf_path}")

    if args.json:
        json_path = os.path.join(args.output, f"{basename}_audit_results.json")
        with open(json_path, 'w') as f:
            json.dump(results, f, indent=2, default=str)
        print(f"    {_c('JSON', '32')}  -> {json_path}")

    total_issues = 0
    severity_totals = Counter()
    for sheet in results['sheets'].values():
        for field in sheet['fields'].values():
            for issue in field['issues']:
                total_issues += 1
                severity_totals[issue['severity']] += 1
        for d in sheet['phantom_duplicates']:
            total_issues += 1
            severity_totals[d['severity']] += 1

    high = severity_totals.get('High', 0)
    med = severity_totals.get('Medium', 0)
    low = severity_totals.get('Low', 0)

    print(f"\n  Audit complete: {_c(str(total_issues) + ' issues found', '1')}")
    print(f"    {_c(f'High: {high}', '31')} | {_c(f'Medium: {med}', '33')} | {_c(f'Low: {low}', '32')}\n")


if __name__ == '__main__':
    main()
