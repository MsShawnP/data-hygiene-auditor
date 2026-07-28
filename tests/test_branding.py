"""Tests that report outputs use the Lailara design system palette."""

from __future__ import annotations

import os
import re
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from audit import generate_excel, generate_html, generate_pdf, run_audit
from data_hygiene_auditor.reporting import palette as P

SAMPLE_PATH = Path(__file__).parent.parent / "samples" / "input" / "sample_messy_data.xlsx"


def _audit_results() -> dict:
    return run_audit(str(SAMPLE_PATH))


class TestHTMLBranding:
    def test_css_uses_palette_canvas(self):
        results = _audit_results()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_html(results, os.path.join(tmpdir, "r.html"))
            html = Path(path).read_text(encoding="utf-8")
            assert P.CANVAS in html
            assert P.CHICAGO_20 in html
            assert P.HONG_KONG_35 in html
            assert P.SINGAPORE_55 in html
            assert P.RED_42 in html

    def test_css_uses_serif_font(self):
        results = _audit_results()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_html(results, os.path.join(tmpdir, "r.html"))
            html = Path(path).read_text(encoding="utf-8")
            assert "Playfair Display" in html
            assert "Source Sans 3" in html

    def test_css_uses_2px_radius(self):
        results = _audit_results()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_html(results, os.path.join(tmpdir, "r.html"))
            html = Path(path).read_text(encoding="utf-8")
            assert P.BORDER_RADIUS in html

    def test_no_old_dark_theme_colors(self):
        results = _audit_results()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_html(results, os.path.join(tmpdir, "r.html"))
            html = Path(path).read_text(encoding="utf-8")
            assert '#1a1a2e' not in html
            assert '#16213e' not in html
            assert '#e94560' not in html


class TestExcelBranding:
    def test_header_fill_is_chicago(self):
        results = _audit_results()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_excel(results, os.path.join(tmpdir, "f.xlsx"))
            wb = load_workbook(path)
            ws = wb["Findings"]
            fill_color = ws.cell(row=1, column=1).fill.fgColor.rgb
            assert fill_color.lower().endswith(P.xl(P.CHICAGO_20))

    def test_header_font_is_source_sans(self):
        results = _audit_results()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_excel(results, os.path.join(tmpdir, "f.xlsx"))
            wb = load_workbook(path)
            ws = wb["Findings"]
            assert ws.cell(row=1, column=1).font.name == P.FONT_SANS_EXCEL

    def test_severity_fills_match_palette(self):
        results = _audit_results()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_excel(results, os.path.join(tmpdir, "f.xlsx"))
            wb = load_workbook(path)
            ws = wb["Findings"]
            sev_colors = set()
            for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
                cell = row[0]
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    sev_colors.add(cell.fill.fgColor.rgb[-6:])
            expected = {P.xl(P.SEV_HIGH_BG), P.xl(P.SEV_MEDIUM_BG), P.xl(P.SEV_LOW_BG)}
            assert sev_colors.issubset(expected | {'000000'})

    def test_no_arial_font(self):
        results = _audit_results()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_excel(results, os.path.join(tmpdir, "f.xlsx"))
            wb = load_workbook(path)
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.font and cell.font.name:
                            assert cell.font.name != "Arial", (
                                f"Cell {cell.coordinate} in '{ws.title}' "
                                f"still uses Arial"
                            )


class TestPDFBranding:
    def test_pdf_generates_successfully(self):
        results = _audit_results()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_pdf(results, os.path.join(tmpdir, "r.pdf"))
            assert os.path.exists(path)
            assert os.path.getsize(path) > 0


class TestPaletteModule:
    # Un-pinned. This class used to assert SEV_HIGH == RED_42, SEV_MEDIUM ==
    # SINGAPORE_55 and SEV_LOW == HONG_KONG_35. palette.py:29-31 defines those
    # names *as* those aliases, so all three assertions were tautologies that
    # could not fail for any palette, correct or not. Worse, the first one
    # asserted the thing the design system forbids: Red-42 is ink -- text and
    # 1px rules -- never a background fill, and reporting/html.py:246,356 plus
    # _SHEET_COLORS/_OVERALL_COLORS fill 23 severity badges, 2 sheet-score
    # chips and 1 filter button with it. pdf.py:127 and excel.py:68 already use
    # the right convention: a Red-95 surface with Red-18 text.
    #
    # Replaced with the assertion the audit asked for. It is strict-xfail, so
    # the suite stays green now and fails the moment the fills are ported,
    # forcing the marker off.
    # Tracked in PLAN.md -- "Red-42 used as a background fill".

    def test_severity_aliases_are_the_documented_palette_steps(self):
        # Not a tautology: these compare against the literal hexes, so a change
        # to either the alias or the underlying step is caught.
        assert P.SEV_MEDIUM == '#ee8a2a'   # Singapore-55
        assert P.SEV_LOW == '#158f75'      # Hong Kong-35

    @pytest.mark.xfail(
        strict=True,
        reason="html.py:246,356 fill severity badges with Red-42, which is ink-only",
    )
    def test_red_42_is_never_a_background_fill_in_the_generated_css(self):
        results = _audit_results()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_html(results, os.path.join(tmpdir, "r.html"))
            css = Path(path).read_text(encoding="utf-8")

        # Fills reach Red-42 through custom properties (--accent, --high), so
        # resolve those first rather than grepping for the raw hex after
        # "background".
        red_vars = {
            name
            for name, value in re.findall(r"--([\w-]+):\s*([^;]+);", css)
            if value.strip().lower() == P.RED_42.lower()
        }
        alternatives = "|".join(re.escape(v) for v in sorted(red_vars)) or r"\0"
        offenders = re.findall(
            rf"(background|background-color|fill)\s*:\s*(?:{re.escape(P.RED_42)}|var\(--(?:{alternatives})\))",
            css,
            re.IGNORECASE,
        )
        assert not offenders, (
            f"Red-42 is used as a fill in {len(offenders)} declaration(s); "
            f"it is ink only. Port the Red-95 surface / Red-18 text convention "
            f"from pdf.py:127 and excel.py:68."
        )

    def test_xl_strips_hash(self):
        assert P.xl('#1f2e7a') == '1f2e7a'
        assert P.xl('1f2e7a') == '1f2e7a'

    def test_all_colors_are_valid_hex(self):
        import re
        hex_pattern = re.compile(r'^#[0-9a-f]{6}$')
        color_attrs = [
            a for a in dir(P)
            if a.isupper() and not a.startswith('FONT') and not a.startswith('BORDER')
        ]
        for attr in color_attrs:
            val = getattr(P, attr)
            if isinstance(val, str) and val.startswith('#'):
                assert hex_pattern.match(val), f"{attr} = {val!r} is not valid hex"
