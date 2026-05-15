"""Tests for fix suggestion generation."""
import os
import tempfile
from pathlib import Path

from data_hygiene_auditor.suggestions import generate_dup_fix, generate_fix


class TestMixedFormatFixes:
    def test_date_fix(self):
        fix = generate_fix('mixed_format', {
            'field_type': 'date',
            'dominant_format': 'YYYY-MM-DD',
        }, 'JoinDate', 'date')
        assert fix is not None
        assert fix['strategy'] == 'normalize_dates'
        assert 'pd.to_datetime' in fix['code']
        assert 'JoinDate' in fix['code']

    def test_phone_fix(self):
        fix = generate_fix('mixed_format', {
            'field_type': 'phone',
            'dominant_format': '(XXX) XXX-XXXX',
        }, 'Phone', 'phone')
        assert fix is not None
        assert fix['strategy'] == 'normalize_phones'
        assert 'Phone' in fix['code']
        assert 'digits' in fix['code']

    def test_currency_fix(self):
        fix = generate_fix('mixed_format', {
            'field_type': 'currency',
            'dominant_format': '$X,XXX.XX',
        }, 'Amount', 'currency')
        assert fix is not None
        assert fix['strategy'] == 'normalize_currency'
        assert 'float' in fix['code']

    def test_unknown_type_returns_none(self):
        fix = generate_fix('mixed_format', {
            'field_type': 'freetext',
        }, 'Notes', 'freetext')
        assert fix is None


class TestPlaceholderFixes:
    def test_placeholder_fix(self):
        fix = generate_fix('placeholder_value', {
            'value': 'N/A',
            'count': 5,
            'pct': 16.7,
        }, 'Email', 'email')
        assert fix is not None
        assert fix['strategy'] == 'replace_placeholders'
        assert 'np.nan' in fix['code']
        assert 'N/A' in fix['code']

    def test_placeholder_type_alias(self):
        fix = generate_fix('placeholder', {
            'value': 'Test',
            'count': 3,
            'pct': 10.0,
        }, 'Name', 'name')
        assert fix is not None


class TestRepetitionFixes:
    def test_repetition_fix(self):
        fix = generate_fix('suspicious_repetition', {
            'value': 'Pending',
            'count': 15,
            'pct': 50.0,
        }, 'Status', 'categorical')
        assert fix is not None
        assert fix['strategy'] == 'flag_repetitions'
        assert 'Pending' in fix['code']
        assert '_review' in fix['code']


class TestWrongPurposeFixes:
    def test_invalid_email(self):
        fix = generate_fix('wrong_purpose', {
            'issue': 'Invalid email format',
            'example': 'not-an-email',
        }, 'Email', 'email')
        assert fix is not None
        assert fix['strategy'] == 'flag_invalid_emails'

    def test_mixed_ids(self):
        fix = generate_fix('wrong_purpose', {
            'issue': 'Mixed ID formats',
            'example': '29 coded vs 1 bare',
        }, 'CustomerID', 'id')
        assert fix is not None
        assert fix['strategy'] == 'standardize_ids'

    def test_text_in_currency(self):
        fix = generate_fix('wrong_purpose', {
            'issue': 'Text in currency field',
            'example': 'free',
        }, 'Price', 'currency')
        assert fix is not None
        assert fix['strategy'] == 'flag_non_numeric'

    def test_generic_wrong_purpose(self):
        fix = generate_fix('wrong_purpose', {
            'issue': 'Numeric value in name field',
            'example': '12345',
        }, 'Name', 'name')
        assert fix is not None
        assert fix['strategy'] == 'flag_misuse'


class TestNullFixes:
    def test_low_missing(self):
        fix = generate_fix('null_analysis', {
            'missing_pct': 10.0,
            'total_missing': 3,
            'total_rows': 30,
        }, 'Name', 'name')
        assert fix is not None
        assert fix['strategy'] == 'fill_missing'
        assert 'mode' in fix['code']

    def test_medium_missing(self):
        fix = generate_fix('null_analysis', {
            'missing_pct': 35.0,
            'total_missing': 7,
            'total_rows': 20,
        }, 'Phone', 'phone')
        assert fix is not None
        assert fix['strategy'] == 'impute_or_flag'
        assert 'Option A' in fix['code']
        assert 'Option B' in fix['code']

    def test_high_missing(self):
        fix = generate_fix('null_analysis', {
            'missing_pct': 80.0,
            'total_missing': 16,
            'total_rows': 20,
        }, 'Fax', 'phone')
        assert fix is not None
        assert fix['strategy'] == 'evaluate_column'
        assert 'drop' in fix['code'].lower()


class TestDuplicateFixes:
    def test_exact_duplicate(self):
        fix = generate_dup_fix('exact_duplicate', {
            'rows': [5, 6],
            'group_size': 2,
        })
        assert fix is not None
        assert fix['strategy'] == 'drop_exact_duplicates'
        assert 'drop_duplicates' in fix['code']

    def test_phantom_duplicate(self):
        fix = generate_dup_fix('phantom_duplicate', {
            'rows': [2, 8, 9],
            'group_size': 3,
        })
        assert fix is not None
        assert fix['strategy'] == 'normalize_and_dedup'
        assert 'strip' in fix['code']

    def test_fuzzy_duplicate(self):
        fix = generate_dup_fix('fuzzy_duplicate', {
            'rows': [3, 7],
            'group_size': 2,
            'match_method': 'fingerprint',
        })
        assert fix is not None
        assert fix['strategy'] == 'review_fuzzy_matches'
        assert '_fuzzy_review' in fix['code']


class TestUnknownType:
    def test_unknown_issue_type(self):
        fix = generate_fix('some_unknown_type', {}, 'col', 'text')
        assert fix is None

    def test_unknown_dup_type(self):
        fix = generate_dup_fix('unknown_dup', {'rows': [1]})
        assert fix is None


class TestFixInReports:
    def test_html_contains_fix_blocks(self):
        from audit import generate_html, run_audit
        sample = (
            Path(__file__).parent.parent
            / "samples" / "input" / "sample_messy_data.xlsx"
        )
        results = run_audit(str(sample))
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_html(
                results, os.path.join(tmpdir, "report.html"),
            )
            content = Path(path).read_text(encoding="utf-8")
            assert "fix-block" in content
            assert "Suggested Fix" in content
            assert "copyFix" in content

    def test_excel_has_fix_column(self):
        from openpyxl import load_workbook

        from audit import generate_excel, run_audit
        sample = (
            Path(__file__).parent.parent
            / "samples" / "input" / "sample_messy_data.xlsx"
        )
        results = run_audit(str(sample))
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_excel(
                results, os.path.join(tmpdir, "findings.xlsx"),
            )
            wb = load_workbook(path)
            ws = wb["Findings"]
            headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
            assert "Suggested Fix" in headers
            has_fix = False
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=9).value
                if val:
                    has_fix = True
                    break
            assert has_fix

    def test_api_fix_suggestions(self):
        from audit import audit_file
        sample = (
            Path(__file__).parent.parent
            / "samples" / "input" / "sample_messy_data.xlsx"
        )
        result = audit_file(str(sample))
        fixes_found = 0
        for s in result.sheets:
            for f in s.findings:
                if f.fix is not None:
                    fixes_found += 1
                    assert f.fix.strategy
                    assert f.fix.description
                    assert f.fix.code
            for d in s.duplicates:
                if d.fix is not None:
                    fixes_found += 1
        assert fixes_found > 0
