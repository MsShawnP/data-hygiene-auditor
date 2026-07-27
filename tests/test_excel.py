"""Behavioral tests for the Excel findings report content."""

from __future__ import annotations

import os
import tempfile

from openpyxl import load_workbook

from audit import generate_excel


def _one_issue_results(issue: dict) -> dict:
    return {
        'input_file': 'x.csv',
        'audit_timestamp': '2024-01-01 00:00:00',
        'overall_score': 50,
        'sheets': {
            'Sheet1': {
                'row_count': 1,
                'col_count': 1,
                'fields': {
                    'Name': {
                        'inferred_type': 'name',
                        'issues': [issue],
                        'profile': {'cardinality': 1, 'uniqueness_pct': 100.0},
                    },
                },
                'phantom_duplicates': [],
                'fuzzy_duplicates': [],
                'schema_violations': [],
                'health_score': 50,
            },
        },
    }


def _findings_row(path: str, field_name: str):
    wb = load_workbook(path)
    ws = wb["Findings"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == field_name:
            return row
    return None


class TestCustomRuleName:
    def test_custom_rule_name_preserved(self):
        # A custom rule's name must appear in both the Issue Type column (D,
        # index 3) and the Description column (F, index 5), not the generic
        # 'custom_rule' / 'Custom Rule' fallbacks.
        results = _one_issue_results({
            'type': 'custom_rule',
            'rule_name': 'No short names',
            'severity': 'Medium',
            'detail': {'message': 'value too short'},
            'why': 'names must be long enough',
        })
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_excel(results, os.path.join(tmpdir, "f.xlsx"))
            row = _findings_row(path, "Name")
        assert row is not None
        assert row[3] == "No short names"          # Issue Type column
        assert "No short names" in str(row[5])      # Description column
        assert row[3] != "custom_rule"


class TestFormulaInjection:
    def test_no_data_cell_is_written_as_a_formula(self):
        # Sheet name, field name, echoed value, and input_file all start with
        # a formula trigger. None may be stored as a live formula; each must
        # be neutralized with a leading apostrophe.
        results = {
            'input_file': '=cmd|calc.csv',
            'audit_timestamp': '2024-01-01 00:00:00',
            'overall_score': 50,
            'sheets': {
                '=Sheet': {
                    'row_count': 1,
                    'col_count': 1,
                    'fields': {
                        '=HYPERLINK("http://evil","x")': {
                            'inferred_type': 'name',
                            'issues': [{
                                'type': 'placeholder',
                                'severity': 'Low',
                                'detail': {'value': '@SUM(A1)', 'count': 1, 'pct': 100},
                                'why': 'w',
                            }],
                            'profile': {'cardinality': 1, 'uniqueness_pct': 100.0},
                        },
                    },
                    'phantom_duplicates': [],
                    'fuzzy_duplicates': [],
                    'schema_violations': [],
                    'health_score': 50,
                },
            },
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            path = generate_excel(results, os.path.join(tmpdir, "f.xlsx"))
            wb = load_workbook(path)
            ws = wb["Findings"]
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    assert cell.data_type != 'f', f"formula leaked: {cell.value!r}"
            assert str(ws.cell(row=2, column=1).value).startswith("'=")   # Sheet
            assert str(ws.cell(row=2, column=2).value).startswith("'=")   # Field
            summary = wb["Summary"]
            assert summary['B4'].data_type != 'f'
            assert str(summary['B4'].value).startswith("'=")
