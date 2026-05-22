"""Excel findings report generator."""

from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..core import describe_issue, describe_schema_violation, issue_example
from . import palette as P


def _write_row(ws, row_num: int, values: list, severity: str,
               thin_border: Border, sev_fills: dict) -> None:
    """Write a styled data row to the findings sheet."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = Font(name=P.FONT_SANS_EXCEL, size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = thin_border
        if col_idx == 5:
            cell.fill = sev_fills.get(severity, PatternFill())


def generate_excel(results: dict[str, Any], output_path: str) -> str:
    """Generate sortable/filterable Excel findings file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"

    headers = [
        "Sheet", "Field", "Inferred Type", "Issue Type", "Severity",
        "Description", "Example / Detail", "Why It Matters",
        "Suggested Fix", "Cardinality", "Uniqueness %",
    ]
    header_font = Font(
        bold=True, color="FFFFFF", size=11, name=P.FONT_SANS_EXCEL,
    )
    header_fill = PatternFill("solid", fgColor=P.xl(P.CHICAGO_20))
    header_align = Alignment(
        horizontal="center", vertical="center", wrap_text=True,
    )
    _border_color = P.xl(P.LONDON_85)
    thin_border = Border(
        left=Side(style='thin', color=_border_color),
        right=Side(style='thin', color=_border_color),
        top=Side(style='thin', color=_border_color),
        bottom=Side(style='thin', color=_border_color),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    sev_fills = {
        'High': PatternFill("solid", fgColor=P.xl(P.SEV_HIGH_BG)),
        'Medium': PatternFill("solid", fgColor=P.xl(P.SEV_MEDIUM_BG)),
        'Low': PatternFill("solid", fgColor=P.xl(P.SEV_LOW_BG)),
    }

    row_num = 2
    for sheet_name, sheet_data in results['sheets'].items():
        for col_name, field_data in sheet_data['fields'].items():
            for issue in field_data['issues']:
                detail = issue['detail']
                itype = issue['type']

                desc = describe_issue(itype, detail)
                example = issue_example(itype, detail, issue)

                fix = issue.get('fix', {})
                fix_text = fix.get('code', '') if fix else ''
                profile = field_data.get('profile', {})
                values = [
                    sheet_name, col_name,
                    field_data['inferred_type'],
                    itype, issue['severity'],
                    desc, example, issue.get('why', ''),
                    fix_text,
                    profile.get('cardinality', ''),
                    profile.get('uniqueness_pct', ''),
                ]
                _write_row(ws, row_num, values, issue['severity'],
                           thin_border, sev_fills)
                row_num += 1

        for dup in sheet_data['phantom_duplicates']:
            dtype = (
                'Exact Duplicate'
                if dup['type'] == 'exact_duplicate'
                else 'Phantom Duplicate'
            )
            desc = (
                f"{dup['group_size']} rows appear to be"
                " the same record"
            )
            example = (
                f"Rows: {', '.join(str(r) for r in dup['rows'])}"
            )
            dup_fix = dup.get('fix', {})
            dup_fix_text = dup_fix.get('code', '') if dup_fix else ''
            values = [
                sheet_name, "(row-level)", "—", dtype,
                dup['severity'], desc, example, dup.get('why', ''),
                dup_fix_text,
            ]
            _write_row(ws, row_num, values, dup['severity'],
                       thin_border, sev_fills)
            row_num += 1

        for fuzz in sheet_data.get('fuzzy_duplicates', []):
            method = fuzz['match_method'].title()
            desc = (
                f"Fuzzy match ({method}):"
                f" {fuzz['group_size']} rows are near-duplicates"
            )
            diffs = fuzz.get('field_differences', {})
            diff_parts = []
            for col, diff in diffs.items():
                if isinstance(diff, dict):
                    vals = diff.get('values', [])
                    diff_parts.append(
                        f"{col}: {', '.join(str(v) for v in vals)}"
                    )
                else:
                    diff_parts.append(
                        f"{col}: {', '.join(str(v) for v in diff)}"
                    )
            example = (
                f"Rows: {', '.join(str(r) for r in fuzz['rows'])}"
            )
            if diff_parts:
                example += f" | Diffs: {'; '.join(diff_parts)}"
            fuzz_fix = fuzz.get('fix', {})
            fuzz_fix_text = (
                fuzz_fix.get('code', '') if fuzz_fix else ''
            )
            values = [
                sheet_name, "(row-level)", "—",
                "Fuzzy Duplicate", fuzz['severity'],
                desc, example, fuzz.get('why', ''),
                fuzz_fix_text,
            ]
            _write_row(ws, row_num, values, fuzz['severity'],
                       thin_border, sev_fills)
            row_num += 1

        for sv in sheet_data.get('schema_violations', []):
            col_name = sv.get('column', '')
            desc = describe_schema_violation(sv)
            example = f"Column: {col_name}"
            values = [
                sheet_name, col_name, "—",
                "Schema Violation", sv['severity'],
                desc, example, sv.get('why', ''), '',
            ]
            _write_row(ws, row_num, values, sv['severity'],
                       thin_border, sev_fills)
            row_num += 1

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 35
    ws.column_dimensions['H'].width = 55
    ws.column_dimensions['I'].width = 50

    ws.auto_filter.ref = f"A1:I{row_num - 1}"
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Summary", 0)
    ws2['A1'] = "Data Hygiene Audit Summary"
    ws2['A1'].font = Font(bold=True, size=14, name=P.FONT_SANS_EXCEL)
    ws2['A3'] = "Health Score:"
    ws2['B3'] = f"{results.get('overall_score', 'N/A')}/100"
    ws2['B3'].font = Font(bold=True, name=P.FONT_SANS_EXCEL, size=12)
    ws2['A4'] = "File:"
    ws2['B4'] = results['input_file']
    ws2['A5'] = "Audit Date:"
    ws2['B5'] = results['audit_timestamp']
    ws2['A6'] = "Total Issues:"
    ws2['B6'] = row_num - 2
    for r in range(3, 7):
        ws2.cell(row=r, column=1).font = Font(
            bold=True, name=P.FONT_SANS_EXCEL, size=10,
        )
        if r != 3:
            ws2.cell(row=r, column=2).font = Font(
                name=P.FONT_SANS_EXCEL, size=10,
            )

    r = 8
    for sname, sdata in results['sheets'].items():
        ws2.cell(row=r, column=1, value=f"Sheet: {sname}")
        ws2.cell(row=r, column=1).font = Font(
            bold=True, name=P.FONT_SANS_EXCEL, size=10,
        )
        ws2.cell(
            row=r, column=2,
            value=f"Score: {sdata.get('health_score', 'N/A')}/100",
        )
        ws2.cell(row=r, column=2).font = Font(name=P.FONT_SANS_EXCEL, size=10)
        r += 1

    trend = results.get('trend')
    if trend:
        r += 1
        ws2.cell(row=r, column=1, value="Trend vs Baseline")
        ws2.cell(row=r, column=1).font = Font(
            bold=True, name=P.FONT_SANS_EXCEL, size=12,
        )
        r += 1
        ws2.cell(row=r, column=1, value="Baseline:")
        ws2.cell(
            row=r, column=2,
            value=f"{trend['baseline_file']} ({trend['baseline_timestamp']})",
        )
        r += 1
        delta = trend['overall_score_delta']
        sign = '+' if delta > 0 else ''
        ws2.cell(row=r, column=1, value="Score Change:")
        ws2.cell(
            row=r, column=2,
            value=f"{trend['overall_score_previous']} → "
            f"{results['overall_score']} ({sign}{delta})",
        )
        r += 1
        td = trend['total_issues_delta']
        sign = '+' if td > 0 else ''
        ws2.cell(row=r, column=1, value="Issues Change:")
        ws2.cell(
            row=r, column=2,
            value=f"{trend['total_issues_previous']} → "
            f"{row_num - 2} ({sign}{td})",
        )
        for rr in range(r - 2, r + 1):
            ws2.cell(row=rr, column=1).font = Font(
                bold=True, name=P.FONT_SANS_EXCEL, size=10,
            )
            ws2.cell(row=rr, column=2).font = Font(
                name=P.FONT_SANS_EXCEL, size=10,
            )

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 40

    wb.save(output_path)
    return output_path
